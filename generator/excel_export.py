"""
Génère le fichier Excel structuré avec 11 onglets.
Chaque onglet a une ligne par commune + une ligne TOTAL/MOYENNE pondérée.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scrapers.insee import compute_zone_totals


# ── Styles ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_TOTAL_FILL  = PatternFill("solid", fgColor="BDD7EE")
_TOTAL_FONT  = Font(bold=True, size=10)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(1, col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _style_row(ws, row_idx: int, is_total: bool = False) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.border = _THIN_BORDER
        cell.alignment = _LEFT
        if is_total:
            cell.font = _TOTAL_FONT
            cell.fill = _TOTAL_FILL


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def _safe_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ".").replace(" ", "").replace("\xa0", ""))
    except ValueError:
        return None


def _sum_col(rows: list[dict], key: str) -> str:
    vals = [_safe_float(r.get(key)) for r in rows]
    vals = [v for v in vals if v is not None]
    return str(round(sum(vals))) if vals else "N/D"


def _weighted_avg(rows: list[dict], key: str, weight_key: str = "pop_2022") -> str:
    total_w = 0.0
    total_v = 0.0
    for r in rows:
        v = _safe_float(r.get(key))
        w = _safe_float(r.get(weight_key))
        if v is not None and w is not None and w > 0:
            total_v += v * w
            total_w += w
    if total_w == 0:
        return "N/D"
    return str(round(total_v / total_w, 1))


def _sort_communes(data: list[dict]) -> list[dict]:
    """Trie les communes par CP puis par nom (les lignes TOTAL restent à la fin)."""
    return sorted(
        [r for r in data if r.get("commune") != "TOTAL ZONE"],
        key=lambda r: (str(r.get("cp", "")), str(r.get("commune", "")).lower()),
    )


# ── Onglets INSEE ─────────────────────────────────────────────────────────────

def _onglet_geo(wb, communes_data: list[dict]) -> None:
    ws = wb.create_sheet("Geo")
    cols = [
        "Commune", "Code INSEE", "CP", "Département", "Région",
        "Population 2022", "Superficie (km²)", "Densité (hab/km²)",
        "Variation pop. 2016-2022 (%)", "Solde naturel (%)", "Solde migratoire (%)",
        "Nb ménages 2022", "Naissances 2022", "Décès 2022", "Statut",
    ]
    _write_header(ws, cols)

    sorted_data = _sort_communes(communes_data)
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""), r.get("code_insee", ""), r.get("cp", ""),
            r.get("departement", ""), r.get("region", ""),
            r.get("pop_2022", ""), r.get("superficie_km2", ""),
            r.get("densite_2022", ""), r.get("var_pop_2016_2022", ""),
            r.get("solde_naturel", ""), r.get("solde_migratoire", ""),
            r.get("nb_menages_2022", ""), r.get("naissances_2022", ""),
            r.get("deces_2022", ""), r.get("_statut", ""),
        ])
        _style_row(ws, row_idx)

    ws.append([
        "TOTAL ZONE", "", "", "", "",
        _sum_col(sorted_data, "pop_2022"),
        _sum_col(sorted_data, "superficie_km2"),
        _weighted_avg(sorted_data, "densite_2022"),
        _weighted_avg(sorted_data, "var_pop_2016_2022"),
        _weighted_avg(sorted_data, "solde_naturel"),
        _weighted_avg(sorted_data, "solde_migratoire"),
        _sum_col(sorted_data, "nb_menages_2022"),
        _sum_col(sorted_data, "naissances_2022"),
        _sum_col(sorted_data, "deces_2022"),
        "Calculé",
    ])
    _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


def _onglet_logement(wb, communes_data: list[dict]) -> None:
    ws = wb.create_sheet("Logement")
    cols = [
        "Commune", "CP",
        "Nb logements 2022",
        "Résidences principales (%)", "Résidences secondaires (%)",
        "Logements vacants (%)", "Ménages propriétaires (%)",
    ]
    _write_header(ws, cols)

    sorted_data = _sort_communes(communes_data)
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""), r.get("cp", ""),
            r.get("nb_logements_2022", ""),
            r.get("part_res_principales_2022", ""),
            r.get("part_res_secondaires_2022", ""),
            r.get("part_logements_vacants_2022", ""),
            r.get("part_proprietaires_2022", ""),
        ])
        _style_row(ws, row_idx)

    ws.append([
        "TOTAL / MOY. PONDÉRÉE", "",
        _sum_col(sorted_data, "nb_logements_2022"),
        _weighted_avg(sorted_data, "part_res_principales_2022"),
        _weighted_avg(sorted_data, "part_res_secondaires_2022"),
        _weighted_avg(sorted_data, "part_logements_vacants_2022"),
        _weighted_avg(sorted_data, "part_proprietaires_2022"),
    ])
    _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


def _onglet_revenus(wb, communes_data: list[dict]) -> None:
    ws = wb.create_sheet("Revenus")
    cols = [
        "Commune", "CP",
        "Nb ménages fiscaux 2021", "Ménages imposés 2021 (%)",
        "Médiane revenu disponible (€)", "Taux de pauvreté (%)",
    ]
    _write_header(ws, cols)

    sorted_data = _sort_communes(communes_data)
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""), r.get("cp", ""),
            r.get("nb_menages_fiscaux_2021", ""),
            r.get("part_menages_imposes_2021", ""),
            r.get("mediane_revenu_2021", ""),
            r.get("taux_pauvrete_2021", ""),
        ])
        _style_row(ws, row_idx)

    ws.append([
        "TOTAL / MOY. PONDÉRÉE", "",
        _sum_col(sorted_data, "nb_menages_fiscaux_2021"),
        _weighted_avg(sorted_data, "part_menages_imposes_2021"),
        _weighted_avg(sorted_data, "mediane_revenu_2021"),
        _weighted_avg(sorted_data, "taux_pauvrete_2021"),
    ])
    _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


def _onglet_emploi(wb, communes_data: list[dict]) -> None:
    ws = wb.create_sheet("Emploi")
    cols = [
        "Commune", "CP",
        "Emploi total 2022", "Part emploi salarié (%)",
        "Variation emploi 2016-2022 (%)",
        "Taux d'activité 15-64 ans (%)",
        "Taux de chômage 15-64 ans (%)",
    ]
    _write_header(ws, cols)

    sorted_data = _sort_communes(communes_data)
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""), r.get("cp", ""),
            r.get("emploi_total_2022", ""),
            r.get("part_emploi_salarie_2022", ""),
            r.get("var_emploi_2016_2022", ""),
            r.get("taux_activite_15_64_2022", ""),
            r.get("taux_chomage_15_64_2022", ""),
        ])
        _style_row(ws, row_idx)

    ws.append([
        "TOTAL / MOY. PONDÉRÉE", "",
        _sum_col(sorted_data, "emploi_total_2022"),
        _weighted_avg(sorted_data, "part_emploi_salarie_2022"),
        _weighted_avg(sorted_data, "var_emploi_2016_2022"),
        _weighted_avg(sorted_data, "taux_activite_15_64_2022"),
        _weighted_avg(sorted_data, "taux_chomage_15_64_2022"),
    ])
    _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


def _onglet_etablissements(wb, communes_data: list[dict]) -> None:
    ws = wb.create_sheet("Etablissements")
    cols = [
        "Commune", "CP",
        "Nb établissements actifs fin 2023",
        "Agriculture (%)", "Industrie (%)", "Construction (%)",
        "Commerce / Transp. / Services (%)", "Admin. / Santé (%)",
        "1 à 9 salariés (%)", "10 salariés ou + (%)",
    ]
    _write_header(ws, cols)

    sorted_data = _sort_communes(communes_data)
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""), r.get("cp", ""),
            r.get("nb_etab_actifs_2023", ""),
            r.get("part_agriculture_2023", ""),
            r.get("part_industrie_2023", ""),
            r.get("part_construction_2023", ""),
            r.get("part_commerce_transp_2023", ""),
            r.get("part_admin_sante_2023", ""),
            r.get("part_etab_1_9_sal_2023", ""),
            r.get("part_etab_10_sal_plus_2023", ""),
        ])
        _style_row(ws, row_idx)

    ws.append([
        "TOTAL / MOY. PONDÉRÉE", "",
        _sum_col(sorted_data, "nb_etab_actifs_2023"),
        _weighted_avg(sorted_data, "part_agriculture_2023"),
        _weighted_avg(sorted_data, "part_industrie_2023"),
        _weighted_avg(sorted_data, "part_construction_2023"),
        _weighted_avg(sorted_data, "part_commerce_transp_2023"),
        _weighted_avg(sorted_data, "part_admin_sante_2023"),
        _weighted_avg(sorted_data, "part_etab_1_9_sal_2023"),
        _weighted_avg(sorted_data, "part_etab_10_sal_plus_2023"),
    ])
    _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


# ── Onglets autres sources ────────────────────────────────────────────────────

def _onglet_pharmacies_mm(wb, pj_data: list[dict]) -> None:
    ws = wb.create_sheet("Pharmacies_MM")
    cols = [
        "Commune", "CP",
        "Nb pharmacies", "Noms pharmacies",
        "Nb magasins mat. médical", "Noms magasins mat. médical",
    ]
    _write_header(ws, cols)

    sorted_data = sorted(pj_data, key=lambda r: (str(r.get("cp", "")), str(r.get("commune", "")).lower()))
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("commune", ""),
            r.get("cp", ""),
            r.get("nb_pharmacies", 0),
            " | ".join(r.get("noms_pharmacies", [])),
            r.get("nb_materiel_medical", 0),
            " | ".join(r.get("noms_materiel_medical", [])),
        ])
        _style_row(ws, row_idx)

    if sorted_data:
        ws.append([
            "TOTAL", "",
            sum(r.get("nb_pharmacies", 0) for r in sorted_data),
            "",
            sum(r.get("nb_materiel_medical", 0) for r in sorted_data),
            "",
        ])
        _style_row(ws, ws.max_row, is_total=True)
    _autofit(ws)


def _onglet_maternites(wb, maternites: list[dict]) -> None:
    ws = wb.create_sheet("Maternites")
    cols = ["Nom", "Ville", "CP", "Statut", "Niveau / Type", "Accouchements/an", "URL source"]
    _write_header(ws, cols)

    sorted_data = sorted(
        maternites,
        key=lambda r: (str(r.get("_cp_recherche", r.get("cp", ""))), str(r.get("nom", "")).lower()),
    )
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("nom", ""),
            r.get("ville", ""),
            r.get("_cp_recherche", r.get("cp", "")),
            r.get("statut", ""),
            r.get("type_niveau", ""),
            r.get("nb_accouchements_an", ""),
            r.get("url_source", ""),
        ])
        _style_row(ws, row_idx)
    _autofit(ws)


def _onglet_sages_femmes(wb, sages_femmes: list[dict]) -> None:
    ws = wb.create_sheet("Sages_Femmes")
    cols = ["Nom", "Prénom", "Codes postaux", "Commune", "Adresse", "Téléphone", "Email"]
    _write_header(ws, cols)

    sorted_data = sorted(sages_femmes, key=lambda r: (str(r.get("nom", "")), str(r.get("prenom", ""))))
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("nom", ""),
            r.get("prenom", ""),
            r.get("code_postaux_display", r.get("cp", "")),
            r.get("commune", ""),
            r.get("adresse", ""),
            r.get("telephone", ""),
            r.get("email", ""),
        ])
        _style_row(ws, row_idx)
    _autofit(ws)


def _onglet_lactariums(wb, lactariums: list[dict]) -> None:
    ws = wb.create_sheet("Lactariums")
    cols = ["Nom", "CP", "Ville", "Adresse", "Téléphone", "Email", "Type", "Don anonyme"]
    _write_header(ws, cols)

    sorted_data = sorted(lactariums, key=lambda r: (str(r.get("cp", "")), str(r.get("nom", "")).lower()))
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("nom", ""),
            r.get("cp", ""),
            r.get("ville", ""),
            r.get("adresse", ""),
            r.get("telephone", ""),
            r.get("email", ""),
            r.get("type", ""),
            r.get("don_anonyme", ""),
        ])
        _style_row(ws, row_idx)
    _autofit(ws)


def _onglet_pmi(wb, pmi: list[dict]) -> None:
    ws = wb.create_sheet("PMI")
    cols = ["Nom", "CP", "Ville", "Adresse", "Téléphone", "Email", "Horaires"]
    _write_header(ws, cols)

    sorted_data = sorted(pmi, key=lambda r: (str(r.get("cp", "")), str(r.get("nom", "")).lower()))
    for row_idx, r in enumerate(sorted_data, 2):
        ws.append([
            r.get("nom", ""),
            r.get("cp", ""),
            r.get("ville", ""),
            r.get("adresse", ""),
            r.get("telephone", ""),
            r.get("email", ""),
            r.get("horaires", ""),
        ])
        _style_row(ws, row_idx)
    _autofit(ws)


def _onglet_synthese(wb, communes_data: list[dict], pj_data: list[dict],
                     maternites: list[dict], lactariums: list[dict],
                     sages_femmes: list[dict], pmi: list[dict],
                     zone_nom: str, dept_code: str, dept_nom: str, region: str) -> None:
    ws = wb.create_sheet("Synthese_Zone")

    t = compute_zone_totals(communes_data)

    def _v(key):
        val = t.get(key)
        return val if val is not None else "N/D"

    rows = [
        ["VARIABLE", "VALEUR ZONE", "TYPE DE CALCUL", "SOURCE / MILLÉSIME"],
        ["Zone", zone_nom, "", "Saisie"],
        ["Département", f"{dept_nom} ({dept_code})", "", "Saisie"],
        ["Région", region, "", "Saisie"],
        ["Nb communes", len([r for r in communes_data if r.get("commune") != "TOTAL ZONE"]), "Comptage", "Calculé"],
        # ── Population ──────────────────────────────────────────────────────
        ["Population totale 2022",              _v("pop_2022"),             "Somme",               "INSEE RP 2022"],
        ["Superficie totale (km²)",             _v("superficie_km2"),       "Somme",               "INSEE RP 2022"],
        ["Densité (hab/km²)",                   _v("densite_2022"),         "Pop totale / Superf.", "INSEE RP 2022"],
        ["Variation pop. 2016-2022 (%)",        _v("var_pop_2016_2022"),    "Moy. pond. / pop.",   "INSEE RP 2022"],
        ["Dont solde naturel (%)",              _v("solde_naturel"),        "Moy. pond. / pop.",   "INSEE RP 2022"],
        ["Dont solde migratoire (%)",           _v("solde_migratoire"),     "Moy. pond. / pop.",   "INSEE RP 2022"],
        ["Nb ménages 2022",                     _v("nb_menages_2022"),      "Somme",               "INSEE RP 2022"],
        ["Naissances domiciliées",              _v("naissances_2022"),      "Somme",               "INSEE État civil"],
        ["Décès domiciliés",                    _v("deces_2022"),           "Somme",               "INSEE État civil"],
        # ── Logement ────────────────────────────────────────────────────────
        ["Nb logements 2022",                   _v("nb_logements_2022"),    "Somme",               "INSEE RP 2022"],
        ["Part résidences principales (%)",     _v("part_res_principales_2022"),  "Moy. pond. / pop.", "INSEE RP 2022"],
        ["Part résidences secondaires (%)",     _v("part_res_secondaires_2022"),  "Moy. pond. / pop.", "INSEE RP 2022"],
        ["Part logements vacants (%)",          _v("part_logements_vacants_2022"),"Moy. pond. / pop.", "INSEE RP 2022"],
        ["Part ménages propriétaires (%)",      _v("part_proprietaires_2022"),    "Moy. pond. / pop.", "INSEE RP 2022"],
        # ── Revenus ─────────────────────────────────────────────────────────
        ["Nb ménages fiscaux",                  _v("nb_menages_fiscaux_2021"),    "Somme",             "INSEE DGI 2021"],
        ["Part ménages imposés (%)",            _v("part_menages_imposes_2021"),  "Moy. pond. / pop.", "INSEE DGI 2021"],
        ["Médiane revenu disponible (€)",       _v("mediane_revenu_2021"),        "Moy. pond. / pop.", "INSEE Filosofi"],
        ["Taux de pauvreté (%)",                _v("taux_pauvrete_2021"),         "Moy. pond. / pop.", "INSEE Filosofi"],
        # ── Emploi ──────────────────────────────────────────────────────────
        ["Emploi total 2022",                   _v("emploi_total_2022"),          "Somme",             "INSEE RP 2022"],
        ["Part emploi salarié (%)",             _v("part_emploi_salarie_2022"),   "Moy. pond. / pop.", "INSEE RP 2022"],
        ["Variation emploi 2016-2022 (%)",      _v("var_emploi_2016_2022"),       "Moy. pond. / pop.", "INSEE RP 2022"],
        ["Taux d'activité 15-64 ans (%)",       _v("taux_activite_15_64_2022"),   "Moy. pond. / pop.", "INSEE RP 2022"],
        ["Taux de chômage 15-64 ans (%)",       _v("taux_chomage_15_64_2022"),    "Moy. pond. / pop.", "INSEE RP 2022"],
        # ── Établissements ──────────────────────────────────────────────────
        ["Nb établissements actifs",            _v("nb_etab_actifs_2023"),        "Somme",             "INSEE REE 2024"],
        ["Part agriculture (%)",                _v("part_agriculture_2023"),      "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part industrie (%)",                  _v("part_industrie_2023"),        "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part construction (%)",               _v("part_construction_2023"),     "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part commerce / transports (%)",      _v("part_commerce_transp_2023"),  "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part admin. / santé (%)",             _v("part_admin_sante_2023"),      "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part étab. 1-9 salariés (%)",         _v("part_etab_1_9_sal_2023"),     "Moy. pond. / pop.", "INSEE REE 2024"],
        ["Part étab. 10+ salariés (%)",         _v("part_etab_10_sal_plus_2023"), "Moy. pond. / pop.", "INSEE REE 2024"],
        # ── Autres sources ──────────────────────────────────────────────────
        ["Nb pharmacies",                       sum(r.get("nb_pharmacies", 0) for r in pj_data),       "Somme", "Pages Jaunes"],
        ["Nb magasins matériel médical",        sum(r.get("nb_materiel_medical", 0) for r in pj_data), "Somme", "Pages Jaunes"],
        ["Nb maternités",                       len([m for m in maternites if m.get("nom") and "Erreur" not in m["nom"]]), "Comptage", "Journal des Femmes"],
        ["Nb lactariums",                       len([l for l in lactariums if "aucun" not in l.get("nom", "").lower()]),   "Comptage", "ALF"],
        ["Nb sages-femmes libérales",           len(sages_femmes),                                     "Comptage (dédoublonné)", "Ordre SF"],
        ["Nb PMI",                              len([p for p in pmi if p.get("nom") and "Erreur" not in p["nom"]]),        "Comptage", "AlloPMI"],
    ]

    for row_idx, row in enumerate(rows, 1):
        ws.append(row)
        for col in range(1, 5):
            cell = ws.cell(row_idx, col)
            cell.border = _THIN_BORDER
            cell.alignment = _LEFT
        if row_idx == 1:
            for col in range(1, 5):
                c = ws.cell(row_idx, col)
                c.font = _HEADER_FONT
                c.fill = _HEADER_FILL
                c.alignment = _CENTER

    ws.freeze_panes = "A2"
    _autofit(ws)


# ── Point d'entrée ────────────────────────────────────────────────────────────

def generate_excel(
    output_path: str,
    zone_nom: str,
    dept_code: str,
    dept_nom: str,
    region: str,
    communes_data: list[dict],
    pj_data: list[dict],
    maternites: list[dict],
    lactariums: list[dict],
    sages_femmes: list[dict],
    pmi: list[dict],
) -> None:
    """Génère le fichier Excel avec 11 onglets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # INSEE (5 onglets thématiques)
    _onglet_geo(wb, communes_data)
    _onglet_logement(wb, communes_data)
    _onglet_revenus(wb, communes_data)
    _onglet_emploi(wb, communes_data)
    _onglet_etablissements(wb, communes_data)
    # Autres sources (4 onglets)
    _onglet_pharmacies_mm(wb, pj_data)
    _onglet_maternites(wb, maternites)
    _onglet_sages_femmes(wb, sages_femmes)
    _onglet_lactariums(wb, lactariums)
    _onglet_pmi(wb, pmi)
    # Synthèse globale (1 onglet)
    _onglet_synthese(wb, communes_data, pj_data, maternites, lactariums, sages_femmes, pmi,
                     zone_nom, dept_code, dept_nom, region)

    wb.save(output_path)
